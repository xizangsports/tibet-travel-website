import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

path = 'G:/CODEX HERMES/西藏体旅官网/酒店供应商总表_综合版.xlsx'
wb = openpyxl.load_workbook(path)
ws_old = wb['2026价格总表']

# ===== Read all data =====
rows_data = []
for r in range(1, ws_old.max_row + 1):
    vals = [ws_old.cell(r, c).value for c in range(1, 11)]
    rows_data.append(vals)

# ===== Parse structured data =====
# Extract: city headers, hotel entries
cities = []  # list of (city_name, [hotel_entries])
current_city = None
current_hotels = []

for r in range(4, len(rows_data)):  # skip title/header rows
    v = rows_data[r]
    city_name = v[0]  # Col A = city header like "📍 拉萨"
    hotel_name = v[1]
    
    if city_name and city_name.startswith('📍'):
        if current_city and current_hotels:
            cities.append((current_city, current_hotels))
        current_city = city_name[2:].strip()  # remove "📍 "
        current_hotels = []
    elif hotel_name:
        current_hotels.append(v)

if current_city and current_hotels:
    cities.append((current_city, current_hotels))

print(f"Found {len(cities)} cities")

# ===== For each hotel, keep only 标间 row =====
def is_standard_room(room_desc):
    """Check if room description is a standard room (标间/双床)"""
    if not room_desc:
        return True  # No description means it's the default/only entry
    desc = str(room_desc)
    keywords_std = ['标间', '双床', '标准', '双人']
    keywords_exclude = ['套房', '单间', '大床房', '大床', '三人间', '亲子', '单间', '榻榻米']
    
    # If it contains standard room keywords AND doesn't contain excluded ones
    has_std = any(k in desc for k in keywords_std)
    has_exclude = any(k in desc for k in keywords_exclude)
    
    if has_std and not has_exclude:
        return True
    if '单/标' in desc or '单标' in desc or '标/单' in desc:
        return True
    if '标间' in desc and '大床' not in desc:
        return True
    return False

def get_hotel_level(star):
    """Return a numeric level for sorting (lower = higher end)"""
    if not star:
        return 90
    s = str(star).strip()
    
    if '豪华五星' in s or '5星' in s or '五星' in s:
        return 10
    if '希尔顿' in s:
        return 15
    if '精品' in s and '5' not in s:
        return 20
    if '准五' in s or '准5星' in s or '挂四准五' in s:
        return 25
    if '4星' in s or '四星' in s or '挂四' in s:
        return 30
    if '准四' in s or '好准四' in s:
        return 35
    if '锦江' in s:
        return 36
    if '3星' in s or '三星' in s or '挂三' in s:
        return 40
    if '挂二' in s or '准三' in s:
        return 45
    if '中高端' in s:
        return 20
    if '哈达集团' in s:
        return 35
    if '维也纳' in s:
        return 30
    if '新中式' in s:
        return 35
    if '帐篷' in s:
        return 60
    if '民宿' in s:
        return 50
    if '经济' in s:
        return 55
    return 50  # default mid range

# ===== Process: keep 标间 only, collect all hotels =====
all_hotels = []  # (city, name, star, room, low, mid, high, phone, addr, price_tier, level_sort)

for city, hotels in cities:
    # Group by hotel name
    hotel_groups = {}
    for h in hotels:
        name = h[1]
        if name not in hotel_groups:
            hotel_groups[name] = []
        hotel_groups[name].append(h)
    
    for name, entries in hotel_groups.items():
        if len(entries) == 1:
            # Single entry - keep as is
            e = entries[0]
            all_hotels.append((city, name, e[2], e[3], e[4], e[5], e[6], e[7], e[8], e[9]))
        else:
            # Multiple room types - keep only 标间
            std_rooms = [e for e in entries if is_standard_room(e[3])]
            if std_rooms:
                e = std_rooms[0]  # Take first match
                all_hotels.append((city, name, e[2], e[3], e[4], e[5], e[6], e[7], e[8], e[9]))
            else:
                # No clear 标间 - take the first entry that's not a suite
                non_suite = [e for e in entries if not any(k in str(e[3] or '') for k in ['套房', '亲子', '三人间', 'LOFT'])]
                if non_suite:
                    e = non_suite[0]
                    all_hotels.append((city, name, e[2], e[3], e[4], e[5], e[6], e[7], e[8], e[9]))
                else:
                    e = entries[0]
                    all_hotels.append((city, name, e[2], e[3], e[4], e[5], e[6], e[7], e[8], e[9]))

print(f"Total hotels with only standard rooms: {len(all_hotels)}")

# ===== Sort: by city order, then by level within city =====
city_order = ['拉萨', '林芝', '那曲', '日喀则', '山南', '阿里', '塔尔钦', '当雄', '萨嘎', '其他地区']
city_rank = {c: i for i, c in enumerate(city_order)}

# For each city, sort hotels by level
sorted_hotels = []
current_city = None
city_bucket = []

for h in all_hotels:
    if h[0] != current_city:
        if city_bucket:
            city_bucket.sort(key=lambda x: get_hotel_level(x[2]))
            sorted_hotels.extend(city_bucket)
        current_city = h[0]
        city_bucket = [h]
    else:
        city_bucket.append(h)

if city_bucket:
    city_bucket.sort(key=lambda x: get_hotel_level(x[2]))
    sorted_hotels.extend(city_bucket)

# ===== Now rebuild the sheet =====
sn = '2026价格总表'
if sn in wb.sheetnames:
    del wb[sn]
ws = wb.create_sheet(sn, 0)  # First position

# Styling
title_font = Font(bold=True, size=14, color='FFFFFF')
header_font = Font(bold=True, size=11, color='FFFFFF')
title_fill = PatternFill('solid', fgColor='1F4E79')
header_fill = PatternFill('solid', fgColor='2E75B6')
city_fill = PatternFill('solid', fgColor='4472C4')
city_font = Font(bold=True, size=12, color='FFFFFF')
green_fill = PatternFill('solid', fgColor='C6EFCE')
gold_fill = PatternFill('solid', fgColor='FFD700')
blue_fill = PatternFill('solid', fgColor='BDD7EE')
gray_fill = PatternFill('solid', fgColor='F2F2F2')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center = Alignment(horizontal='center', vertical='center')
wrap = Alignment(vertical='center', wrap_text=True)

# Row 1: Title
ws.merge_cells('A1:J1')
ws['A1'] = '🏔️ 西藏体育旅行社·酒店供应商总表（2026年综合版）— 标间价格·按档次排列'
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = Alignment(horizontal='center')

# Row 2: Description
ws.merge_cells('A2:J2')
ws['A2'] = '整合来源：U盘I:\\酒店(53个文件)+桌面西藏常用酒店(全部文件) | 绿色=2026年最新价格 | 每个酒店仅保留标间价格行 | 按档次高→低排列'
ws['A2'].font = Font(italic=True, size=10, color='555555')

# Row 3: Headers
headers = ['城市', '酒店名称', '档次', '房型说明', '淡季(元)', '平季(元)', '旺季(元)', '联系电话', '地址/备注', '价格区间']
for c, h in enumerate(headers, 1):
    cell = ws.cell(3, c, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

# ===== Write data =====
row = 4
prev_city = None

for h in sorted_hotels:
    city, name, star, room, low, mid, high, phone, addr, tier = h
    
    # City header row
    if city != prev_city:
        ws.merge_cells(f'A{row}:J{row}')
        cell = ws.cell(row, 1, f'📍 {city}')
        cell.font = city_font
        cell.fill = city_fill
        for c in range(1, 11):
            ws.cell(row, c).border = thin_border
        row += 1
        prev_city = city
    
    # Data row
    vals = [None, name, star, room, low, mid, high, phone, addr, tier]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row, c, v)
        cell.border = thin_border
        cell.alignment = wrap if c in (4, 8, 9) else center if c in (1, 3, 5, 6, 7, 10) else Alignment(vertical='center')
    
    # Level-based coloring for column C (档次)
    level_cell = ws.cell(row, 3)
    lvl = str(star or '')
    if '豪华五星' in lvl or '5星' in lvl or '五星' in lvl:
        level_cell.fill = gold_fill
        level_cell.font = Font(bold=True)
    elif '准五' in lvl or '挂四准五' in lvl or '准5' in lvl:
        level_cell.fill = gold_fill
        level_cell.font = Font(bold=False)
    elif '希尔顿' in lvl:
        level_cell.fill = gold_fill
    elif '精品' in lvl:
        level_cell.fill = PatternFill('solid', fgColor='FFE699')
    elif '4星' in lvl or '四星' in lvl or '挂四' in lvl:
        level_cell.fill = blue_fill
    elif '维也纳' in lvl:
        level_cell.fill = blue_fill
    elif '准四' in lvl or '好准四' in lvl:
        level_cell.fill = PatternFill('solid', fgColor='D9E2F3')
    elif '3星' in lvl or '三星' in lvl or '挂三' in lvl:
        level_cell.fill = gray_fill
    elif '锦江' in lvl:
        level_cell.fill = PatternFill('solid', fgColor='E2EFDA')
    elif '帐篷' in lvl:
        level_cell.fill = PatternFill('solid', fgColor='FCE4D6')
    
    # Green highlight for cells with 2026 prices (non-empty price columns)
    has_price = any(v and v != '—' and v != '' for v in [low, mid, high])
    if has_price:
        for c in [5, 6, 7]:
            ws.cell(row, c).fill = green_fill
    
    row += 1

# Column widths
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 14
ws.column_dimensions['H'].width = 26
ws.column_dimensions['I'].width = 40
ws.column_dimensions['J'].width = 18

# Save to temp location first, then copy
import shutil, os
tmp = 'C:/Users/Administrator/Desktop/酒店供应商总表_综合版_TEMP.xlsx'
wb.save(tmp)
# Try to replace the original
try:
    os.remove(path)
except:
    pass
shutil.move(tmp, path)
print(f'✅ 2026价格总表已重构：{len(sorted_hotels)}家酒店，每个仅保留标间行，按档次排列')
